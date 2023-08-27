import requests
from datetime import date
from datetime import datetime
import time
import pandas as pd
import csv
from tempfile import NamedTemporaryFile
import shutil
import os
import openpyxl
from openpyxl import Workbook

def div0(x,y):
    try:
        return round(x/y,2)
    except ZeroDivisionError:
        return 0

def try_fetch(field):
    try:
        return field
    except KeyError:
        return 'NaN'

def notify(title, text):
        os.system("""
                  osascript -e 'display notification "{}" with title "{}"'
                  """.format(text, title))

def get_todays_games():
    tournament_code= {'LaLiga': 8, 'Premier League': 17, 'Championship': 18, 'EFL Cup': 17, 'UEFA Champions League' : 7, 'UEFA Europa League': 679, 'Serie A': 23} 
    #print(list(tournament_code.keys()))
    #rabis za player_stats preveri API ce se kaj spreminja - to je stevila v GET seasons pod user_count!
    league_season_code = {'LaLiga': 52376, 'Premier League': 52186, 'Championship': 52367, 'UEFA Champions League' : 41897, 'UEFA Europa League': 44509, 'Serie A': 52760}  
    #glej insomnia: Unique season IDs
    today          = date.today()
    today_formated = today.strftime("%Y-%m-%d")
    today_slo_form = today.strftime("%d.%m.%Y")
    with open('/Users/Roho11/Desktop/pyscripts/check_game_date.txt', 'r') as file:
        last_read_date = file.readline()
        file.close()
    if last_read_date < today_formated:
        with open('/Users/Roho11/Desktop/pyscripts/check_game_date.txt', 'w') as file:
            file.write(today_formated)
            file.close()
        url            = f'https://api.sofascore.com/api/v1/sport/football/scheduled-events/{today}'
        payload        = ""
        headers        = {
            "authority": "api.sofascore.com",
            "accept": "*/*",
            "accept-language": "en-GB,en-US;q=0.9,en;q=0.8",
            "cache-control": "max-age=0",
            "if-none-match": """W/"dfff6baed4""",
            "origin": "https://www.sofascore.com",
            "referer": "https://www.sofascore.com/",
            "sec-ch-ua": """Not?A_Brand";v="8", "Chromium";v="108", "Google Chrome";v="108""",
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": "macOS",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-site",
            "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36"
                 }
        response       = requests.request("GET", url, data=payload, headers=headers)
        all_games_today= response.json()
        games_in_play  = []
        game_count     = 0
        for game in all_games_today['events']:
            dict_tekme = {}
            prvenstvo = game['tournament']['name']
            drz_prv   = game['tournament']['category']['name']
            tip       = game['status']['type']
            timestamp  =game['startTimestamp']
            strtimstmp = datetime.fromtimestamp(timestamp)
            starttime  =strtimstmp.strftime('%H:%M')
            starttime1 =strtimstmp.strftime('%H%M')
            startdate  =strtimstmp.strftime('%d.%m.%Y')
            #if prvenstvo in ['LaLiga', 'Premier League', 'Championship', 'UEFA Champions League', 'UEFA Europa League', 'Serie A'] and drz_prv in ['Spain', 'England', 'Europe', 'Italy'] and tip not in ['finished', 'postponed'] and startdate == today_slo_form:
            if prvenstvo in list(tournament_code.keys()) and drz_prv in ['Spain', 'England', 'Europe', 'Italy'] and tip not in ['finished', 'postponed'] and startdate == today_slo_form:


                game_count += 1
                opis       =game['status']['description']
                gameslug   =game['slug']
                hometeam_id=game['homeTeam']['id']
                hometeam   =game['homeTeam']['name']
                awayteam_id=game['awayTeam']['id']
                awayteam   =game['awayTeam']['name']
                print(hometeam, awayteam)
                id_tekme   =game['id'] #rabis v linki postav tekme
                custom_id  =game['customId'] #rabis v linki detaili tekme
                #timestamp  =game['startTimestamp']
                #strtimstmp = datetime.fromtimestamp(timestamp)
                #starttime  =strtimstmp.strftime('%H:%M')
                #starttime1 =strtimstmp.strftime('%H%M')
                #startdate  =strtimstmp.strftime('%d.%m.%Y')
                dict_tekme['id'] = id_tekme
                dict_tekme['Custom id'] = custom_id
                dict_tekme['Slug'] = gameslug
                dict_tekme['Tournament'] = prvenstvo
                dict_tekme['HTid'] = hometeam_id
                dict_tekme['HomeTeam'] = hometeam
                dict_tekme['ATid'] = awayteam_id
                dict_tekme['AwayTeam'] = awayteam
                dict_tekme['Status'] = opis
                dict_tekme['Start time'] = starttime
                dict_tekme['Date'] = startdate   
                dict_tekme['ConfLineups'] = 'Not checked'
                url = f"https://api.sofascore.com/api/v1/event/{id_tekme}"
                payload = ""
                headers = {
                    "authority": "api.sofascore.com",
                    "accept": "*/*",
                    "accept-language": "sl-SI,sl;q=0.9",
                    "cache-control": "max-age=0",
                    "if-none-match": """W/"550acc26db""",
                    "origin": "https://www.sofascore.com",
                    "referer": "https://www.sofascore.com/",
                    "sec-ch-ua": """Not?A_Brand";v="8", "Chromium";v="108", "Brave";v="108""",
                    "sec-ch-ua-mobile": "?0",
                    "sec-ch-ua-platform": """macOS""",
                    "sec-fetch-dest": "empty",
                    "sec-fetch-mode": "cors",
                    "sec-fetch-site": "same-site",
                    "sec-gpc": "1",
                    "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36"
                            }
                response  = requests.request("GET", url, data=payload, headers=headers)
                game_data = response.json()
                try: 
                    ref = game_data['event']['referee']
                    ref_name = game_data['event']['referee']['name']
                    ref_yc   = game_data['event']['referee']['yellowCards']
                    ref_rc   = game_data['event']['referee']['redCards']
                    ref_yrc  = game_data['event']['referee']['yellowRedCards']
                    ref_gms  = game_data['event']['referee']['games']
                    refypg   = div0(ref_yc,ref_gms)
                    refrpg   = div0(ref_rc,ref_gms)
                except KeyError:
                    ref_name = 'NaN'
                    ref_yc = 0
                    ref_rc = 0
                    ref_yrc = 0
                    ref_gms = 0
                    refypg = 0
                    refrpg = 0
                dict_tekme['Referee']      = ref_name
                dict_tekme['CPG']          = round(refypg+refrpg,2)
                dict_tekme['YellowCardPG'] = refypg
                dict_tekme['RedCardPG']    = refrpg
                #time.sleep(2)
                games_in_play.append(dict_tekme)
        gip_df         = pd.DataFrame(games_in_play)
        try:
            gip_df.sort_values(by=['Tournament','Date','Start time', 'CPG'],
                    ascending=[True, False, False, False])
            gip_df.to_csv(f'/Users/Roho11/Desktop/pyscripts/GamesInPlay/{today_formated} Games in Play.csv')
            #gip_df.to_excel(f'/Users/Roho11/Desktop/pyscripts/GamesInPlay/{today_formated} Games in Play.xlsx')
            notify('Games in Play',f'{game_count} game(s) stored in {today_formated} Games in Play.xlsx')
            os.makedirs(f'/Users/Roho11/Desktop/pyscripts/GamesInPlay/{today_formated}')
        except KeyError:
            print("KeyError: No games in selected tournaments.")
    else:
        print("Today's games already scraped.")
        print("Checking for lineups...")
        #preveri postave
        lineup_csv_count = 0
        tempfile = NamedTemporaryFile(mode='w', delete=False) 
        fields = ['row_id','id','Custom id','Slug','Tournament','HTid','HomeTeam','ATid','AwayTeam','Status','Start time','Date','ConfLineups','Referee','CPG', 'YellowCardPG','RedCardPG']
        dictionary=dict(zip(fields, fields))
        with open(f'/Users/Roho11/Desktop/pyscripts/GamesInPlay/{today_formated} Games in Play.csv') as f, tempfile:
            reader = csv.DictReader(f, fieldnames=fields)
            #next(reader, None) # ne vzame header v for loop
            writer = csv.DictWriter(tempfile, fieldnames=fields)
            #writer.writerow(dictionary)
            for row in reader:
                ht        = row['HomeTeam']
                at        = row['AwayTeam']
                game_id   = row['id']
                starttime = row['Start time']
                print(starttime)
                if starttime == 'Start time': #prvo vrstico vzame (header)
                    st_clean = 'Start time'
                else:
                    st_clean  = datetime.strptime(str(starttime),'%H:%M').strftime('%H%M')
                    ref       = row['Referee']
                    print(st_clean)

                if row['ConfLineups'] in ['Not checked', 'Not confirmed']: 
                    url = f'https://api.sofascore.com/api/v1/event/{game_id}/lineups'
                    payload = ""
                    headers = {
                            "authority": "api.sofascore.com",
                            "accept": "*/*",
                            "accept-language": "en-GB,en-US;q=0.9,en;q=0.8",
                            "cache-control": "max-age=0",
                            "origin": "https://www.sofascore.com",
                            "referer": "https://www.sofascore.com/",
                            "sec-ch-ua": """Not?A_Brand";v="8", "Chromium";v="108", "Google Chrome";v="108""",
                            "sec-ch-ua-mobile": "?0",
                            "sec-ch-ua-platform": """macOS""",
                            "sec-fetch-dest": "empty",
                            "sec-fetch-mode": "cors",
                            "sec-fetch-site": "same-site",
                            "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36"
                              }
                    response = requests.request("GET", url, data=payload, headers=headers)
                    game_data     = response.json()
                    print(game_data)
                    try:
                        lineups = game_data['confirmed']
                        #notify(f'{ht} : {at}',f'Lineups confirmed: {lineups}')
                        if game_data['confirmed'] != True:
                            #update status
                            updated_row = {'row_id': row['row_id'],'id': row['id'],'Custom id': row['Custom id'],'Slug': row['Slug'], 'Tournament': row['Tournament'],'HTid': row['HTid'], 'HomeTeam': row['HomeTeam'],'ATid': row['ATid'],'AwayTeam': row['AwayTeam'],'Status': row['Status'],'Start time': row['Start time'],'Date': row['Date'],'ConfLineups': 'Not confirmed','Referee': row['Referee'],'CPG': row['CPG'],'YellowCardPG': row['YellowCardPG'],'RedCardPG': row['RedCardPG']}
                            writer.writerow(updated_row)
                        else:
                            #update status confirmed
                            updated_row = {'row_id': row['row_id'],'id': row['id'],'Custom id': row['Custom id'],'Slug': row['Slug'], 'Tournament': row['Tournament'],'HTid': row['HTid'], 'HomeTeam': row['HomeTeam'],'ATid': row['ATid'],'AwayTeam': row['AwayTeam'],'Status': row['Status'],'Start time': row['Start time'],'Date': row['Date'],'ConfLineups': 'Confirmed','Referee': row['Referee'],'CPG': row['CPG'],'YellowCardPG': row['YellowCardPG'],'RedCardPG': row['RedCardPG']} 
                            writer.writerow(updated_row)
                            
                            #ustvari CSV z postavami posameznih tekem
                            
                            notify('Confirmed lineups', f'{ht} - {at} lineups added.')
                            game_lineup      = []
                            lineup_csv_count += 1
                            for player in game_data['home']['players']:
                                player_dict = {}
                                player_id   = player['player']['id']
                                player_name = player['player']['shortName']
                                player_pos  = player['player']['position']
                                team_id     = row['HTid']
                                tournament  = row['Tournament']
                                tour_id     = tournament_code[tournament]
                                season_id   = league_season_code[tournament]
                                if player['substitute'] == True:
                                    player_xi = "Sub"
                                else:
                                    player_xi = "XI"
                                url = f"https://api.sofascore.com/api/v1/player/{player_id}/unique-tournament/{tour_id}/season/{season_id}/statistics/overall"
                                payload = ""
                                headers = {
                                    "authority": "api.sofascore.com",
                                    "accept": "*/*",
                                    "accept-language": "sl-SI,sl;q=0.8",
                                    "cache-control": "max-age=0",
                                    "origin": "https://www.sofascore.com",
                                    "referer": "https://www.sofascore.com/",
                                    "sec-ch-ua": '''Not_A Brand";v="99", "Brave";v="109", "Chromium";v="109''',
                                    "sec-ch-ua-mobile": "?0",
                                    "sec-ch-ua-platform": '''macOS''',
                                    "sec-fetch-dest": "empty",
                                    "sec-fetch-mode": "cors",
                                    "sec-fetch-site": "same-site",
                                    "sec-gpc": "1",
                                    "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36"
                                }
                                response = requests.request("GET", url, data=payload, headers=headers)
                                player_stats = response.json()
                                #ce igralec nima statistike vrne error 404
                                try:
                                    if player_stats['error']['code'] == 404:
                                        print(f'{ht}: {player_id} {player_name} nima statistike')
                                        player_dict['Team']           = row['HomeTeam']
                                        player_dict['Starter']        = player_xi
                                        player_dict['Player']         = player_name
                                        player_dict['Pos']            = player_pos
                                        player_dict['apps']           = 'NaN'
                                        player_dict['mPG']            = 'NaN'
                                        player_dict['yC']             = 'NaN'
                                        player_dict['dRC']            = 'NaN'
                                        player_dict['CardPG']         = 'NaN'
                                        player_dict['TacklesPG']      = 'NaN'
                                        player_dict['tacklesWP']      = 'NaN'
                                        player_dict['foulPG']         = 'NaN'
                                        player_dict['fouledPG']       = 'NaN'
                                        player_dict['succDrblP']      = 'NaN'
                                        game_lineup.append(player_dict)
                                except KeyError:
                                    player_dict['Team']           = row['HomeTeam']
                                    player_dict['Starter']        = player_xi
                                    player_dict['Player']         = player_name
                                    player_dict['Pos']            = player_pos
                                    player_dict['apps']           = player_stats['statistics']['appearances']
                                    min_played                    = player_stats['statistics']['minutesPlayed']
                                    apps                          = player_stats['statistics']['appearances']
                                    player_dict['mPG']            = div0(min_played,apps)
                                    player_dict['yC']             = player_stats['statistics']['yellowCards']
                                    player_dict['dRC']            = player_stats['statistics']['directRedCards']
                                    cards                         = player_stats['statistics']['yellowCards'] + player_stats['statistics']['directRedCards']
                                    player_dict['CardPG']         = div0(cards, apps)
                                    tackles                       = player_stats['statistics']['tackles']
                                    player_dict['TacklesPG']      = div0(tackles,apps)
                                    player_dict['tacklesWP']      = str(round(player_stats['statistics']['tacklesWonPercentage'])) + ' %'
                                    fouls                         = player_stats['statistics']['fouls']
                                    player_dict['foulPG']         = div0(fouls,apps)
                                    wasFouled                     = player_stats['statistics']['wasFouled']
                                    player_dict['fouledPG']       = div0(wasFouled,apps)
                                    player_dict['succDrblP']      = str(round(player_stats['statistics']['successfulDribblesPercentage']))+' %'
                                    game_lineup.append(player_dict)
                            for player in game_data['away']['players']:
                                player_dict = {}
                                player_id   = player['player']['id']
                                player_name = player['player']['shortName']
                                try:
                                    player_pos  = player['player']['position']
                                except KeyError:
                                    player_pos = 'NaN'
                                team_id     = row['ATid']
                                tournament  = row['Tournament']
                                tour_id     = tournament_code[tournament]
                                season_id   = league_season_code[tournament]
                                if player['substitute'] == True:
                                    player_xi = "Sub"
                                else:
                                    player_xi = "XI"
                                url = f"https://api.sofascore.com/api/v1/player/{player_id}/unique-tournament/{tour_id}/season/{season_id}/statistics/overall"
                                payload = ""
                                headers = {
                                    "authority": "api.sofascore.com",
                                    "accept": "*/*",
                                    "accept-language": "sl-SI,sl;q=0.8",
                                    "cache-control": "max-age=0",
                                    "origin": "https://www.sofascore.com",
                                    "referer": "https://www.sofascore.com/",
                                    "sec-ch-ua": '''Not_A Brand";v="99", "Brave";v="109", "Chromium";v="109''',
                                    "sec-ch-ua-mobile": "?0",
                                    "sec-ch-ua-platform": '''macOS''',
                                    "sec-fetch-dest": "empty",
                                    "sec-fetch-mode": "cors",
                                    "sec-fetch-site": "same-site",
                                    "sec-gpc": "1",
                                    "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36"
                                }
                                response = requests.request("GET", url, data=payload, headers=headers)
                                player_stats = response.json()
                                #print(f'{player_name}: {player_stats}')
                                #ce igralec nima statistike vrne error 404
                                try:
                                    if player_stats['error']['code'] == 404:
                                        
                                        print(f'{at}: {player_id} {player_name} nima statistike')
                                        player_dict['Team']           = row['AwayTeam']
                                        player_dict['Starter']        = player_xi
                                        player_dict['Player']         = player_name
                                        player_dict['Pos']            = player_pos
                                        player_dict['apps']           = 'NaN'
                                        player_dict['mPG']            = 'NaN'
                                        player_dict['yC']             = 'NaN'
                                        player_dict['dRC']            = 'NaN'
                                        player_dict['CardPG']         = 'NaN'
                                        player_dict['TacklesPG']      = 'NaN'
                                        player_dict['tacklesWP']      = 'NaN'
                                        player_dict['foulPG']         = 'NaN'
                                        player_dict['fouledPG']       = 'NaN'
                                        player_dict['succDrblP']      = 'NaN'
                                        game_lineup.append(player_dict)
                                except KeyError:
                                    player_dict['Team']           = row['AwayTeam']
                                    player_dict['Starter']        = player_xi
                                    player_dict['Player']         = player_name
                                    player_dict['Pos']            = player_pos
                                    player_dict['apps']           = player_stats['statistics']['appearances']
                                    min_played                    = player_stats['statistics']['minutesPlayed']
                                    apps                          = player_stats['statistics']['appearances']
                                    player_dict['mPG']            = div0(min_played,apps)
                                    player_dict['yC']             = player_stats['statistics']['yellowCards']
                                    player_dict['dRC']            = player_stats['statistics']['directRedCards']
                                    cards                         = player_stats['statistics']['yellowCards'] + player_stats['statistics']['directRedCards']
                                    player_dict['CardPG']         = div0(cards, apps)
                                    tackles                       = player_stats['statistics']['tackles']
                                    player_dict['TacklesPG']      = div0(tackles,apps)
                                    player_dict['tacklesWP']      = str(round(player_stats['statistics']['tacklesWonPercentage'])) + ' %'
                                    fouls                         = player_stats['statistics']['fouls']
                                    player_dict['foulPG']         = div0(fouls,apps)
                                    wasFouled                     = player_stats['statistics']['wasFouled']
                                    player_dict['fouledPG']       = div0(wasFouled,apps)
                                    player_dict['succDrblP']      = str(round(player_stats['statistics']['successfulDribblesPercentage'])) + ' %'
                                    game_lineup.append(player_dict)
                                    
                            ps_df = pd.DataFrame(game_lineup)
                            #ps_df.to_csv(f'/Users/Roho11/Desktop/pyscripts/GamesInPlay/{today_formated}/{starttime} {ht} - {at}.csv')
                            ps_df.to_excel(f'/Users/Roho11/Desktop/pyscripts/GamesInPlay/{today_formated}/{st_clean} {ht} vs {at}.xlsx')
                            
                            #book = openpyxl.load_workbook(f'/Users/Roho11/Desktop/pyscripts/GamesInPlay/{today_formated}/{starttime} {ht} - {at}.xlsx')
                            #sheet = book.active

                            
                            #for row in sheet.iter_rows(min_row=1, max_col=6, max_row=sheet.max_row, values_only=True):
                            #    if row[5] > 75:
                            #        # If the value is greater than 75, set the font style to bold
                            #        cell = sheet.cell(row=row[0], column=6)
                            #        cell.font = openpyxl.styles.Font(bold=True)
                           
                            #book.save(f'/Users/Roho11/Desktop/pyscripts/GamesInPlay/{today_formated}/{starttime} {ht} - {at}.xlsx')
                        
                            
                            
                    except KeyError:
                        print("Verjetno se ni niti predvidenih postav... calma")     
                else:
                    print(f'No changes on game {game_id}! or header row.')
                    updated_row = {'row_id': row['row_id'],'id': row['id'],'Custom id': row['Custom id'],'Slug': row['Slug'], 'Tournament': row['Tournament'],'HTid': row['HTid'], 'HomeTeam': row['HomeTeam'],'ATid': row['ATid'],'AwayTeam': row['AwayTeam'],'Status': row['Status'],'Start time': row['Start time'],'Date': row['Date'],'ConfLineups': row['ConfLineups'],'Referee': row['Referee'],'CPG': row['CPG'], 'YellowCardPG': row['YellowCardPG'],'RedCardPG': row['RedCardPG']}
                    writer.writerow(updated_row)
        shutil.move(tempfile.name, f'/Users/Roho11/Desktop/pyscripts/GamesInPlay/{today_formated} Games in Play.csv')

get_todays_games()